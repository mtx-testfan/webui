'''
读取excel的方式
数据结构：列表套字典
'''
# 导包
import os
import openpyxl
class ReadExcel:
    def get_excel(self,head):
        # wb-工作簿
        # 成功excel文件 然后顺利读取xlsx文件
        # wb = openpyxl.load_workbook('../data/data.xlsx')
        # 加r,\\,/ ,以下路径表示方法是绝对路径,缺点：灵活性差  框架：架构一样的 ，相同的
        # 绝对路径：     base_dir                        +/data/ + 'data.xlsx'
        # os.sep 自动去匹配你的系统的/,\(自动区分mac还是windows)
        #  D:/company/auto/20200606/lesson6/lesson6_2/=》如何获取这个路径？？？？
        base_dir = os.path.dirname(os.path.dirname(__file__))
        wb = openpyxl.load_workbook(base_dir+'%sdata%s' % (os.sep,os.sep)+'data.xlsx')

        # print(wb)
        # ws -sheet页
        ws = wb['测试用例']
        # data = [
          #第一个字典：username：第2行,第1列的cell值;
        #                   password：第2行，第2列的cell值
        #         {'username': 'yaoyao', 'password': 'yaoyao1'},
        # 第二个字典：username：第3行, 第1列的cell值;
        #         password：第3行，第2列的cell值
        #         {'username': 'beihe', 'password': 'beihe'},
        # 第三个字典：username：第4行, 第1列的cell值;
        #         password：第4行，第2列的cell值
        #         {'username': 'yaoyao1', 'password': 'yaoyao6'},
        #
        #         ]
        # 行数怎么确定，->最大行 从第二行开始，到最大行 for i in range()
        # 列数怎么确定->head里面的元素个数决定有几列
        all_cases=[]
        # 控制行数
        for row in range(2,ws.max_row+1):
            dict = {}
            # 控制列数的，head里面有多少个元素，那么就有几列
            col = 1
            # 控制列数的
            for key in head:
                # {'username':第二行第一列的值}
                dict[key] = ws.cell(row,col).value
                col += 1
            all_cases.append(dict)
        # print(all_cases)
        return all_cases



if __name__ == '__main__':
    all_cases = ReadExcel().get_excel(['username','password'])
    print(all_cases)

